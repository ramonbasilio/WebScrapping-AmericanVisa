from selenium import webdriver
from datetime import datetime
from openpyxl import load_workbook
from os.path import exists

import openpyxl
import os
import time
import dictionaryDaysMonths
import datetime

WEBSITE_PAGE = "https://ais.usvisa-info.com/pt-br/niv/users/sign_in"


class AmericanVisaAccess():
    def __init__(self, email, password):
        self.email = email
        self.password = password
        self.controlWhile = True
        self.totalDays = 0
        self.availableDays = []
        self.firstDayOf2022 = datetime.datetime(2022, 1, 1)

        self.browser = webdriver.Chrome()
        self.browser.set_window_position(1024, 600)
        self.browser.maximize_window()

        self.browser.get(WEBSITE_PAGE)
        time.sleep(1)
        self.browser.find_element_by_id('user_email').send_keys(self.email)
        time.sleep(1)
        self.browser.find_element_by_id(
            'user_password').send_keys(self.password)
        time.sleep(1)
        self.browser.find_element_by_xpath(
            '/html/body/div[5]/main/div[3]/div/div[1]/div/form/div[3]/label/div').click()
        time.sleep(1)
        self.browser.find_element_by_xpath(
            '//*[@id="new_user"]/p[1]/input').click()
        time.sleep(1)
        self.browser.find_element_by_xpath(
            '//*[@id="main"]/div[2]/div[3]/div[1]/div/div/div[1]/div[2]/ul/li/a').click()
        time.sleep(1)
        self.acordeonList = self.browser.find_elements_by_class_name(
            'accordion-item')
        self.acordeonList[3].click()
        time.sleep(1)
        self.browser.find_element_by_xpath(
            '/html/body/div[4]/main/div[2]/div[2]/div/section/ul/li[4]/div/div/div[2]/p[2]/a').click()  # tratar essa parte
        time.sleep(1)
        self.browser.find_element_by_xpath(
            '/html/body/div[4]/main/div[4]/div/div/form/fieldset[1]/ol/fieldset/div/div[2]/div[3]/li[1]/input').click()
        time.sleep(1)

        while(self.controlWhile):
            self.auxList = []
            self.currentYear = self.browser.find_element_by_xpath(
                '/html/body/div[5]/div[1]/div/div/span[2]').text
            # lista com todos os nomes das classes (tags)
            self.dayList = self.browser.find_elements_by_class_name(
                'ui-state-default')
            self.currentMonth = self.browser.find_element_by_xpath(
                f'/html/body/div[5]/div[{1}]/div/div/span[1]').text
            self.amountDaysOfCurrentMonth = dictionaryDaysMonths.dic[self.currentMonth]

            for i in range(self.amountDaysOfCurrentMonth):
                self.auxList.append(self.dayList[i].tag_name)

            if 'a' in self.auxList:
                for i in range(self.amountDaysOfCurrentMonth):
                    if self.auxList[i] == 'a':
                        print(
                            f'Avaliable date: {i+1}/{self.currentMonth}/{self.currentYear}')
                        self.availableDays.append(str(i+1))
                        self.totalDays = + self.totalDays + 1
                print(f'Total Days: {self.totalDays}')
                self.firstClosestDate = f'{self.availableDays[0]}/{self.currentMonth}/{self.currentYear}'
                print(self.firstClosestDate)
                self.controlWhile = False
                self.availableDays = ', '.join(self.availableDays)

            else:
                self.browser.find_element_by_xpath(
                    '/html/body/div[5]/div[2]/div/a/span').click()

    def checkFileIsCreated(self):
        if os.path.exists('SheetData.xlsx') == False:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws["A1"] = "Search Index"
            self.ws["B1"] = "Search Date"
            self.ws["C1"] = "Month"
            self.ws["D1"] = "Year"
            self.ws["E1"] = "Number of days available"
            self.ws["F1"] = "Days available"
            self.ws["G1"] = "First Closest Date"
            self.ws["H1"] = "Proximity Parameter"
            self.wb.save(filename='SheetData.xlsx')
        else:
            self.wb = load_workbook("SheetData.xlsx")
            self.ws = self.wb.active  # ativa
            self.ws["A1"] = "Search Index"
            self.ws["B1"] = "Search Date"
            self.ws["C1"] = "Month"
            self.ws["D1"] = "Year"
            self.ws["E1"] = "Number of days available"
            self.ws["F1"] = "Days available"
            self.ws["G1"] = "First Closest Date"
            self.ws["H1"] = "Proximity Parameter"
            self.wb.save(filename="SheetData.xlsx")

    def seeNextIndex(self):
        self.wb = load_workbook("SheetData.xlsx")
        self.ws = self.wb.active  # ativa
        cont = 1
        a = True
        while a:
            if self.ws["A"][0].value == None:
                try:
                    self.ws["A"][1]
                except:
                    a = False
                    cont = 1
                    return cont
            try:
                self.ws["A"][cont]
            except:
                a = False
                return cont + 1
            cont = + cont + 1

    def proximityParameter(self):
        self.dateFinal = datetime.datetime.strptime(
            self.firstClosestDate, '%d/%B/%Y')
        self.dateInicial = datetime.datetime(2022, 1, 1)
        self.proximityParameterValue = str(
            self.dateFinal - self.dateInicial)[:3]
        return self.proximityParameterValue
