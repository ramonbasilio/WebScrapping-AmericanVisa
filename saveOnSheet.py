from datetime import datetime
from openpyxl import load_workbook

class save():
    def __init__(self,
                 index,
                 currentMonth,
                 currentYear,
                 totalDays,
                 availableDays,
                 firstClosestDate,
                 proximityParameter):

        self.index = index
        self.currentDate = str(datetime.now())[:16]
        self.currentMonth = currentMonth
        self.currentYear = currentYear
        self.totalDays = totalDays
        self.availableDays = availableDays
        self.firstClosestDate = firstClosestDate
        self.proximityParameter = proximityParameter
        self.wb = load_workbook("SheetData.xlsx")
        self.ws = self.wb.active 

        self.ws[f"A{self.index}"] = self.index
        self.ws[f"B{self.index}"] = self.currentDate
        self.ws[f"C{self.index}"] = self.currentMonth
        self.ws[f"D{self.index}"] = self.currentYear
        self.ws[f"E{self.index}"] = self.totalDays
        self.ws[f"F{self.index}"] = self.availableDays
        self.ws[f"G{self.index}"] = self.firstClosestDate
        self.ws[f"H{self.index}"] = self.proximityParameter
        self.wb.save(filename="SheetData.xlsx")
