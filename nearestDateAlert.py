from openpyxl import load_workbook


class alert():
    def __init__(self, index, proximityParameter):
        self.index = index
        self.proximityParameter = proximityParameter

    def nearestDateAlert(self):
        self.wb = load_workbook("SheetData.xlsx")
        self.ws = self.wb.active  # ativa
        lastValue = self.ws[f"H{self.index-1}"].value
        if lastValue > self.proximityParameter:
            return "There's a closer date!"
        if lastValue == self.proximityParameter:
            return "Same as last date."
        else:
            return "No forecast!"
